from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from database import get_db
from models import Lesson, Student, Tutor
from schemas import LessonCreate, LessonUpdate, LessonResponse
from pydantic import BaseModel
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from datetime import date as date_type
import io

router = APIRouter(prefix="/lessons", tags=["lessons"])

@router.get("/", response_model=list[LessonResponse])
def get_lessons(db: Session = Depends(get_db)):
    return db.query(Lesson).all()


@router.get("/export")
def export_lessons(
    student_id: int | None = Query(default=None),
    tutor_id: int | None = Query(default=None),
    date_from: date_type | None = Query(default=None),
    date_to: date_type | None = Query(default=None),
    db: Session = Depends(get_db)
):
    query = db.query(Lesson)
    if student_id:
        query = query.filter(Lesson.student_id == student_id)
    if tutor_id:
        query = query.filter(Lesson.tutor_id == tutor_id)
    if date_from:
        query = query.filter(Lesson.date >= date_from)
    if date_to:
        query = query.filter(Lesson.date <= date_to)
    lessons = query.order_by(Lesson.date).all()

    student_map = {s.id: s for s in db.query(Student).filter(Student.id.in_({l.student_id for l in lessons})).all()}
    tutor_map = {t.id: t for t in db.query(Tutor).filter(Tutor.id.in_({l.tutor_id for l in lessons})).all()}

    wb = load_workbook('lessons_4_1_26.xlsx')
    sheet_name = date_type.today().strftime('%Y-%m-%d')
    ws = wb.create_sheet(title=sheet_name)

    ws.append(['Date', 'WeekStart', 'DayOfWeek', 'Day', 'Student', 'Hrs', 'Fee', 'Tutor', 'Pay', 'PayStatus', 'Notes'])

    for i, lesson in enumerate(lessons, start=2):
        student = student_map.get(lesson.student_id)
        tutor = tutor_map.get(lesson.tutor_id)
        student_name = f'{student.first_name} {student.last_name[0]}' if student else ''
        tutor_name = f'{tutor.first_name} {tutor.last_name}' if tutor else ''
        ws.append([
            lesson.date,
            f'=A{i}-WEEKDAY(A{i},2)+1',
            f'=WEEKDAY(A{i},2)',
            f'=TEXT(A{i},"dddd")',
            student_name,
            lesson.hrs,
            lesson.fee,
            tutor_name,
            lesson.tutor_payout,
            'x' if lesson.pay_status else '',
            lesson.notes,
        ])

    if lessons:
        tbl = Table(displayName=f'Export{sheet_name.replace("-", "")}', ref=f'A1:K{len(lessons) + 1}')
        ws.add_table(tbl)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="lessons_{sheet_name}.xlsx"'}
    )


@router.get("/{lesson_id:int}", response_model=LessonResponse)
def get_lesson(lesson_id: int, db: Session = Depends(get_db)):
    lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
    if not lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    return lesson

def handle_fee_and_payout(new_lesson: Lesson, lesson_in: LessonCreate | LessonUpdate, db_student: Student, db_tutor: Tutor):
    # override fee if provided
    if lesson_in.fee_override is not None:
        new_lesson.fee = lesson_in.fee_override
        new_lesson.is_fee_overridden = True
    # or calculate it based on normal rate and hours if not
    else:
        new_lesson.fee = (lesson_in.hrs or 0) * db_student.rate
        new_lesson.is_fee_overridden = False

    # Tutor payout logic:
    # fee_override does not affect tutor payout for normal lessons (protects tutor from discounts on the lesson,
    # and protects business from paying tutor for more time than worked if extra fee charged): hrs x rate.
    # Exception: cancelled lessons (hrs=0) with a penalty fee — tutor gets a proportional cut of the penalty.
    # To override in any case, set tutor_pay_override explicitly.
    if lesson_in.tutor_pay_override is not None:
        new_lesson.tutor_payout = lesson_in.tutor_pay_override
        new_lesson.is_tutor_payout_overridden = True
    elif lesson_in.hrs == 0 and lesson_in.fee_override is not None:
        new_lesson.tutor_payout = lesson_in.fee_override * (db_tutor.pay_rate / db_student.rate)
        new_lesson.is_tutor_payout_overridden = False
    else:
        new_lesson.tutor_payout = (lesson_in.hrs or 0) * db_tutor.pay_rate
        new_lesson.is_tutor_payout_overridden = False

@router.post("/", response_model=LessonResponse, status_code=201)
def create_lesson(lesson_in: LessonCreate, db: Session = Depends(get_db)):
    db_student = db.query(Student).filter(Student.id == lesson_in.student_id).first()
    db_tutor = db.query(Tutor).filter(Tutor.id == lesson_in.tutor_id).first()

    if not db_student:
        raise HTTPException(status_code=404, detail="Student not found")
    if not db_tutor:
        raise HTTPException(status_code=404, detail="Tutor not found")

    new_lesson = Lesson(**lesson_in.model_dump(exclude={"fee_override", "tutor_pay_override"}))
    handle_fee_and_payout(new_lesson, lesson_in, db_student, db_tutor)

    db.add(new_lesson)
    db.commit()
    db.refresh(new_lesson)
    return new_lesson


@router.post("/bulk_create", response_model=list[LessonResponse], status_code=201)
def bulk_create_lessons(lessons_in: list[LessonCreate], db: Session = Depends(get_db)):
    # 2 queries, as opposed to looping per lesson and querying for student and tutor each time
    student_ids = {lesson.student_id for lesson in lessons_in}
    tutor_ids = {lesson.tutor_id for lesson in lessons_in}

    students = {s.id: s for s in db.query(Student).filter(Student.id.in_(student_ids)).all()}
    tutors = {t.id: t for t in db.query(Tutor).filter(Tutor.id.in_(tutor_ids)).all()}

    new_lessons = []
    for lesson_in in lessons_in:
        if lesson_in.student_id not in students:
            raise HTTPException(status_code=404, detail=f"Student with id {lesson_in.student_id} not found")
        if lesson_in.tutor_id not in tutors:
            raise HTTPException(status_code=404, detail=f"Tutor with id {lesson_in.tutor_id} not found")
        db_student = students[lesson_in.student_id]
        db_tutor = tutors[lesson_in.tutor_id]
        new_lesson = Lesson(**lesson_in.model_dump(exclude={"fee_override", "tutor_pay_override"}))
        handle_fee_and_payout(new_lesson, lesson_in, db_student, db_tutor)
        new_lessons.append(new_lesson)
    
    db.add_all(new_lessons)
    db.commit()
    for lesson in new_lessons:
        db.refresh(lesson)
    return new_lessons
    
    
@router.put("/{lesson_id:int}", response_model=LessonResponse)
def update_lesson(lesson_id: int, lesson_in: LessonUpdate, db: Session = Depends(get_db)):
    db_lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
    if not db_lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")

    db_student = db.query(Student).filter(Student.id == db_lesson.student_id).first()
    db_tutor = db.query(Tutor).filter(Tutor.id == db_lesson.tutor_id).first()

    # Update non-derived fields first
    for key, value in lesson_in.model_dump(exclude={"fee_override", "tutor_pay_override"}).items():
        setattr(db_lesson, key, value)

    handle_fee_and_payout(db_lesson, lesson_in, db_student, db_tutor)

    db.commit()
    db.refresh(db_lesson)
    return db_lesson


@router.delete("/{lesson_id:int}", response_model=LessonResponse)
def delete_lesson(lesson_id: int, db: Session = Depends(get_db)):
    db_lesson = db.query(Lesson).filter(Lesson.id == lesson_id).first()
    if not db_lesson:
        raise HTTPException(status_code=404, detail="Lesson not found")
    db.delete(db_lesson)
    db.commit()
    return db_lesson

class BulkDeleteRequest(BaseModel):
    lesson_ids: list[int]

@router.delete("/bulk_delete")
def bulk_delete_lessons(request: BulkDeleteRequest, db: Session = Depends(get_db)):
    lessons = db.query(Lesson).filter(Lesson.id.in_(request.lesson_ids)).all()
    if not lessons:
        raise HTTPException(status_code=404, detail="No matching lessons found")
    for lesson in lessons:
        db.delete(lesson)
    db.commit()
    return {"deleted": len(lessons)}